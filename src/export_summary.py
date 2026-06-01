import os
import numpy as np
import pandas as pd
from scipy.signal import find_peaks, savgol_filter
from data_parser import parse_wafer_data
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================================
# [설정] 아웃라이어(Outlier) 및 에러 판단 기준
# ==========================================================
THRESHOLD = {
    'IL_MIN': -20.0,
    'ER_MIN': 10.0,
    'VPIL_MIN': 0.2,
    'VPIL_MAX': 3.0
}

zip_path = "../dat/HY202103"

# 저장 디렉토리 분리
save_dir_xlsx = "../res/xlsx"
save_dir_csv = "../res/csv"
os.makedirs(save_dir_xlsx, exist_ok=True)
os.makedirs(save_dir_csv, exist_ok=True)

target_wafers = ['D07', 'D08', 'D23', 'D24']
L_length = 0.05


def q_sub(x, y):
    if len(x) < 3: return x[np.argmin(y)]
    idx = np.argmin(y)
    if idx == 0 or idx == len(x) - 1: return x[idx]
    c = np.polyfit(x[idx - 1:idx + 2], y[idx - 1:idx + 2], 2)
    return -c[1] / (2 * c[0]) if abs(c[0]) > 1e-12 else x[idx]


def check_status(row):
    reasons = []
    if pd.isna(row['IL_dB']):
        reasons.append("IL 데이터 없음")
    elif row['IL_dB'] < THRESHOLD['IL_MIN']:
        reasons.append(f"IL 손실 과다(<{THRESHOLD['IL_MIN']})")

    if pd.isna(row['ER_dB']):
        reasons.append("ER 데이터 없음")
    elif row['ER_dB'] < THRESHOLD['ER_MIN']:
        reasons.append(f"ER 낮음(<{THRESHOLD['ER_MIN']})")

    if pd.isna(row['VpiL_Vcm']):
        reasons.append("VpiL 계산 실패")
    elif row['VpiL_Vcm'] < THRESHOLD['VPIL_MIN'] or row['VpiL_Vcm'] > THRESHOLD['VPIL_MAX']:
        reasons.append(f"VpiL 범위 초과({THRESHOLD['VPIL_MIN']}~{THRESHOLD['VPIL_MAX']})")

    return ("정상", "") if not reasons else ("이상 발생", ", ".join(reasons))


# ==========================================================
# 엑셀 서식 자동 적용 함수
# ==========================================================
def apply_excel_style(worksheet, dataframe):
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    error_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    link_font = Font(color='0563C1', underline='single')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                    bottom=Side(style='thin'))

    col_idx = {col: i for i, col in enumerate(dataframe.columns, 1)}

    for col_num, column_title in enumerate(dataframe.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

        for row_num in range(2, len(dataframe) + 2):
            data_cell = worksheet.cell(row=row_num, column=col_num)
            data_cell.border = border

            # 1. Folder_Link 열 (개별 다이 통합 PNG 열기)
            if column_title == 'Folder_Link':
                target_image = data_cell.value
                if target_image:
                    data_cell.value = f'=HYPERLINK("{target_image}", "🖼️ 개별 다이 이미지 보기")'
                    data_cell.font = link_font
                    data_cell.alignment = Alignment(horizontal='center')

            # 2. 웨이퍼 맵 하이퍼링크 (웨이퍼 맵 통합 이미지 열기)
            elif column_title == 'Map_Image_Link':
                target_image = data_cell.value
                if target_image:
                    data_cell.value = f'=HYPERLINK("{target_image}", "🗺️ 웨이퍼 맵 통합 이미지 보기")'
                    data_cell.font = link_font
                    data_cell.alignment = Alignment(horizontal='center')

            # 3. 박스 플롯 하이퍼링크 (박스 플롯 통합 이미지 열기)
            elif column_title == 'BoxPlot_Image_Link':
                target_image = data_cell.value
                if target_image:
                    data_cell.value = f'=HYPERLINK("{target_image}", "📊 박스 플롯 통합 이미지 보기")'
                    data_cell.font = link_font
                    data_cell.alignment = Alignment(horizontal='center')

            else:
                if column_title in ['IL_dB', 'ER_dB', 'VpiL_Vcm'] and isinstance(data_cell.value, (int, float)):
                    data_cell.number_format = '0.000'
                elif column_title == 'R_Squared' and isinstance(data_cell.value, (int, float)):
                    data_cell.number_format = '0.0000'

            # 에러 행 하이라이트
            if 'Status' in col_idx:
                status_value = worksheet.cell(row=row_num, column=col_idx['Status']).value
                if status_value == "이상 발생" and column_title not in ['Folder_Link', 'Map_Image_Link', 'BoxPlot_Image_Link']:
                    data_cell.fill = error_fill

    for col in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                text = str(cell.value)
                max_length = max(max_length, sum(2 if ord(c) > 127 else 1 for c in text))
        worksheet.column_dimensions[column_letter].width = min(max_length + 3, 50)


# ==========================================================
# 메인 실행 블록
# ==========================================================
print("🚀 데이터 분석 및 계산을 시작합니다...")
data_list = []

for d in parse_wafer_data(zip_path, target_wafers):
    wafer, band, c, r = d['wafer_id'], d['band'], d['die_c'], d['die_r']
    radius = np.sqrt(c ** 2 + r ** 2)
    date_str = d.get('date', 'Unknown_Date')

    m = (d['ref_data']['wl'] >= d['wl_min']) & (d['ref_data']['wl'] <= d['wl_max'])
    v_ref_wl, v_ref_il = d['ref_data']['wl'][m], d['ref_data']['il'][m]
    if len(v_ref_wl) < 31: continue

    sm_ref = savgol_filter(v_ref_il, 31, 3)
    poly_func = np.poly1d(np.polyfit(v_ref_wl, sm_ref, 3))
    y_fitted = poly_func(v_ref_wl)

    ss_res = np.sum((sm_ref - y_fitted) ** 2)
    ss_tot = np.sum((sm_ref - np.mean(sm_ref)) ** 2)
    r_squared = 1.0 if ss_tot == 0 else 1 - (ss_res / ss_tot)

    max_peak = float('-inf')
    for b in d['bias_data_list']:
        if b['bias'] is None: continue
        mb = (b['wl'] >= d['wl_min']) & (b['wl'] <= d['wl_max'])
        if np.any(mb):
            max_peak = max(max_peak, np.max(b['il'][mb]))
    il_val = max_peak if max_peak != float('-inf') else np.nan

    max_er = 0.0
    for b in d['bias_data_list']:
        if b['bias'] is None: continue
        mb = (b['wl'] >= d['wl_min']) & (b['wl'] <= d['wl_max'])
        v_wl, v_il = b['wl'][mb], b['il'][mb]
        if len(v_wl) == 0: continue
        flat_il = v_il - poly_func(v_wl)
        peaks, _ = find_peaks(flat_il, prominence=3.0, distance=30)
        final_il = flat_il - np.poly1d(np.polyfit(v_wl[peaks], flat_il[peaks], 1))(v_wl) if len(peaks) >= 2 else flat_il
        max_er = max(max_er, np.percentile(final_il, 99) - np.percentile(final_il, 1))
    er_val = max_er if max_er > 0 else np.nan

    vpil_val = np.nan
    z_data = next((b for b in d['bias_data_list'] if b['bias'] == 0.0), None)
    if z_data:
        m0 = (z_data['wl'] >= d['wl_min']) & (z_data['wl'] <= d['wl_max'])
        w0, i0 = z_data['wl'][m0], z_data['il'][m0]
        v0, _ = find_peaks(-(savgol_filter(i0, 31, 3) - poly_func(w0)), prominence=0.3, distance=20)
        if len(v0) >= 2:
            fsr, cwl = np.mean(np.diff(w0[v0])), w0[v0[np.argmin(np.abs(w0[v0] - d['target_wl']))]]
            volts, p_pi = [], []
            for b in d['bias_data_list']:
                if b['bias'] is None: continue
                w, i = b['wl'], b['il']
                mb = (w >= d['wl_min']) & (w <= d['wl_max'])
                wb, ib = w[mb], i[mb]
                mloc = (wb >= cwl - fsr / 2.5) & (wb <= cwl + fsr / 2.5)
                if np.sum(mloc) >= 5:
                    volts.append(b['bias'])
                    p_pi.append(
                        2.0 * (q_sub(wb[mloc], savgol_filter(ib[mloc], 11, 3) - poly_func(wb[mloc])) - cwl) / fsr)
            if len(volts) >= 5:
                vpil_0V = L_length / max(np.abs(np.polyder(np.poly1d(np.polyfit(volts, p_pi, 2)))(0.0)), 1e-5)
                if 0.1 <= vpil_0V <= 10.0: vpil_val = vpil_0V

    data_list.append({
        'Wafer': wafer, 'Band': band, 'Date': date_str, 'Column': c, 'Row': r, 'Radius': radius,
        'IL_dB': il_val, 'ER_dB': er_val, 'VpiL_Vcm': vpil_val, 'R_Squared': r_squared
    })

# --- 전체 데이터 정리 및 날짜 병합 ---
df_full = pd.DataFrame(data_list).groupby(['Wafer', 'Band', 'Date', 'Column', 'Row', 'Radius'], as_index=False).min(
    numeric_only=True)
df_full['Status'], df_full['Reason'] = zip(*df_full.apply(check_status, axis=1))


# [추가된 부분] 0603, 0604 데이터 병합
def merge_midnight_dates(date_val):
    date_str = str(date_val)
    if '0603' in date_str or '0604' in date_str:
        return '20190603'
    return date_str


df_full['Date'] = df_full['Date'].apply(merge_midnight_dates)
print("🕒 0603 및 0604 날짜 데이터를 '0603_0604_Combined'로 통합했습니다.")


# 🌟 1. 개별 다이 이미지 경로 생성
def generate_die_paths(row):
    wafer = str(row['Wafer'])
    date = str(row['Date'])
    c = int(row['Column'])
    r = int(row['Row'])
    band = str(row['Band'])
    base_dir = f"C:\\Users\\sodlg\\PycharmProjects\\PE2\\res\\png\\{wafer}\\{date}"
    image_name = f"HY202103_{wafer}_({c},{r})_LION1_DCM_{band}.png"
    return f"{base_dir}\\{image_name}"


df_full['Folder_Link'] = df_full.apply(generate_die_paths, axis=1)

print("--------------------------------------------------")
print("▶ 결과 파일 저장을 시작합니다...")

# ==========================================================
# [순수 CSV 파일 생성] (링크 컬럼 제거됨, 합본 포함)
# ==========================================================
csv_df = df_full.drop(columns=['Folder_Link'], errors='ignore')

for wafer_id in csv_df['Wafer'].unique():
    wafer_csv = csv_df[csv_df['Wafer'] == wafer_id]
    csv_file_path = os.path.join(save_dir_csv, f"{wafer_id}_Process_result.csv")
    wafer_csv.to_csv(csv_file_path, index=False, encoding='utf-8-sig')
    print(f"  - 개별 순수 CSV 저장 완료: {csv_file_path}")

total_csv_path = os.path.join(save_dir_csv, "Total_Process_result.csv")
csv_df.to_csv(total_csv_path, index=False, encoding='utf-8-sig')
print(f"  - 🌟 전체 통합 CSV 저장 완료: {total_csv_path}")

print("--------------------------------------------------")

# ==========================================================
# [엑셀 XLSX 파일 생성] (병합 이미지 링크 포함)
# ==========================================================
for wafer_id in df_full['Wafer'].unique():
    wafer_df = df_full[df_full['Wafer'] == wafer_id].copy()
    wafer_file_path = os.path.join(save_dir_xlsx, f"{wafer_id}_Process_result.xlsx")
    with pd.ExcelWriter(wafer_file_path, engine='openpyxl') as writer:
        wafer_df.to_excel(writer, index=False, sheet_name='Analysis_Result')
        apply_excel_style(writer.sheets['Analysis_Result'], wafer_df)
    print(f"  - 개별 XLSX 저장 완료: {wafer_file_path}")

total_xlsx_path = os.path.join(save_dir_xlsx, "Total_Process_result.xlsx")
with pd.ExcelWriter(total_xlsx_path, engine='openpyxl') as writer:
    df_full.to_excel(writer, index=False, sheet_name='Analysis_Result')
    apply_excel_style(writer.sheets['Analysis_Result'], df_full)
print(f"  - 🌟 전체 통합 XLSX 저장 완료: {total_xlsx_path}")

# ==========================================================
# 3. 새로운 Analysis.xlsm 및 Analysis.csv 생성 (통합 이미지 직접 링크)
# ==========================================================
df_index = df_full[['Wafer', 'Date', 'Band']].drop_duplicates().reset_index(drop=True)


# 🌟 웨이퍼 맵 통합 파일 경로 링크 생성
def generate_map_paths(row):
    wafer = str(row['Wafer'])
    date = str(row['Date'])
    band = str(row['Band'])
    return f"C:\\Users\\sodlg\\PycharmProjects\\PE2\\res\\png\\WaferMap\\{wafer}\\{date}\\Summary_WaferMap_{wafer}_{band}_{date}.png"


# 🌟 박스 플롯 통합 파일 경로 링크 생성
def generate_boxplot_paths(row):
    wafer = str(row['Wafer'])
    date = str(row['Date'])
    band = str(row['Band'])
    return f"C:\\Users\\sodlg\\PycharmProjects\\PE2\\res\\png\\BoxPlot\\{wafer}\\{date}\\Summary_BoxPlot_{wafer}_{band}_{date}.png"


# 링크 적용 (이름 변경)
df_index['Map_Image_Link'] = df_index.apply(generate_map_paths, axis=1)
df_index['BoxPlot_Image_Link'] = df_index.apply(generate_boxplot_paths, axis=1)

# Analysis.csv 저장 (경로 제거)
analysis_csv_path = os.path.join(save_dir_csv, "Analysis.csv")
df_index.drop(columns=['Map_Image_Link', 'BoxPlot_Image_Link'], errors='ignore').to_csv(analysis_csv_path,
                                                                                          index=False,
                                                                                          encoding='utf-8-sig')
print(f"  - 🌟 마스터 CSV 저장 완료: {analysis_csv_path}")

# Analysis.xlsm 저장 (대시보드 형태)
total_file_path = os.path.join(save_dir_xlsx, "Analysis.xlsm")
with pd.ExcelWriter(total_file_path, engine='openpyxl') as writer:
    df_index.to_excel(writer, index=False, sheet_name='Dashboard_Links')
    apply_excel_style(writer.sheets['Dashboard_Links'], df_index)
print(f"  - 🌟 마스터 XLSX 저장 완료 (통합 이미지 다이렉트 링크 연결 완료): {total_file_path}")

print("--------------------------------------------------")
print(f"✅ 모든 분석 및 파일 생성이 성공적으로 완료되었습니다!")